import openpyxl
import re
import collections
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import pandas as pd

import sys



try:
    file = sys.argv[1]
    document = openpyxl.load_workbook(file)

    #sheet is a workbook
    sheet = document.active

    def containsLPU(string):
        return bool((re.search('LPU',string,re.IGNORECASE)))

    def getEdges(sheet):
        return ('A1:'+str(get_column_letter(sheet.max_column))+str(sheet.max_row))

    def getHeaders(sheet):
        headings = []
        for row in sheet[1]:
            headings.append(row.value)
        return(headings)

    getHeaders(sheet)

    rowsContainsLPU = list()

    for row in sheet.iter_rows():
        for column in range(len(row)):
            try:
                if (containsLPU(str(row[column].value))):
                    if row not in rowsContainsLPU:
                        rowsContainsLPU.append(row)
            except TypeError:
                pass

    def writeIntoNewSheet(raw):

        def applystyle(sheet):
            #apply table
            tab = Table(displayName="Table1", ref=getEdges(sheet))

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            ws2.add_table(tab)
        
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active

        mr = len(raw)
        mc = len(raw[0])

        #add headers
        for column in range(mc):
                ws2.cell(row=1, column = column+1).value = getHeaders(sheet)[column]
                ws2.cell(row=1, column = column+1).font = Font(color='000000', italic=True)

        for i in range (mr): 
            for j in range (mc): 
                # reading cell value from source excel file 
                c = raw[i][j]
                
                # writing the read value to destination excel file 
                ws2.cell(row = i+2, column = j+1).value = c.value
        
        applystyle(ws2)
        
        # saving the destination excel file 
        wb2.save(filename='LPU_extracted.xlsx')

    writeIntoNewSheet(rowsContainsLPU)
except FileNotFoundError as not_found:
    print(not_found)